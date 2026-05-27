# -*- coding: utf-8 -*-
"""
CRM Generator — Shikardos Team
Генерирует Excel-файл с 7 листами из данных Instagram, YouTube, Telegram.

Запуск:
  pip install openpyxl
  python create_crm.py

CSV-файл CRM_Instagram_DeepScan_442.csv должен лежать рядом со скриптом
(или в той же папке).
"""

import csv
import os
import re
import sys
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl не установлен. Устанавливаю...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ============================================================
# КОНФИГУРАЦИЯ
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(SCRIPT_DIR, f"CRM_Shikardos_Team_{datetime.now().strftime('%Y-%m-%d')}.xlsx")

# Ищем CSV в нескольких местах
_possible_csv_paths = [
    os.path.join(SCRIPT_DIR, "CRM_Instagram_DeepScan_442.csv"),
    os.path.expandvars(r"%APPDATA%\Claude\local-agent-mode-sessions\f01f13a9-f550-4493-be26-3d902532cba8\2d8b64eb-6bf4-4c02-b2c3-828d6de61155\local_be2278f6-93e0-4162-b64c-5a0d2a48b3a2\uploads\CRM_Instagram_DeepScan_442.csv"),
    os.path.join(os.path.expanduser("~"), "Downloads", "CRM_Instagram_DeepScan_442.csv"),
    os.path.join(os.path.expanduser("~"), "Desktop", "CRM_Instagram_DeepScan_442.csv"),
]
IG_CSV = None
for _p in _possible_csv_paths:
    if os.path.exists(_p):
        IG_CSV = _p
        break
if IG_CSV is None:
    IG_CSV = _possible_csv_paths[0]  # fallback для сообщения об ошибке

# Дубликаты Instagram (username в нижнем регистре)
DUPLICATES = {
    "arina_alexx", "sebastien", "klymenko_bogdan", "alex_slobozhenko",
    "olegtorbosov", "vlad_iss_love", "misterr_alex", "margo_bylinina",
    "thomaskralowru", "metamorfozu", "viola.instalova", "nikrider",
    "valya_denisov", "nixxionx", "harkushao", "serditov", "maks_filantrop",
    "alex_agafonov__", "davitube_", "teenismyage", "gipsy.mba"
}

# ============================================================
# ВСТРОЕННЫЕ ДАННЫЕ: YouTube (95 каналов)
# ============================================================
YOUTUBE_DATA = [
    ["AI Прорыв","@VolchenkoAI",29200,"США","","AI/Автоматизация"],
    ["AQTRAFF","@aqtraff",6890,"Россия","TG,IG","Арбитраж"],
    ["ArbitrageScanner","@ArbitrageScanner",12000,"США","TG","Крипто/Арбитраж"],
    ["Анна AI VISIONS","@ai.visionsss",2780,"","TG","AI/Нейросети"],
    ["Big Money","@BigMoneylive",799000,"Украина","TG,IG,FB","Бизнес/Интервью"],
    ["Coin Post","@CoinPost",312000,"Латвия","TG","Крипто"],
    ["Do4a","@DO4AMEDIA",133000,"Россия","TG,IG,VK","Фитнес/Лайфстайл"],
    ["fxalexg","@fxalexg__",1240000,"США","","Трейдинг/Финансы"],
    ["G GATE","@weareggate",4970,"Казахстан","TG,IG","Affiliate/iGaming"],
    ["Gambling.pro","@Gamblingpro",23600,"","","Арбитраж/Gambling"],
    ["HOT Protocol","@hot-labs",1260000,"США","TG","Web3/Крипто"],
    ["IKIGAI","@IKIGAIMAIN",406000,"Россия","TG","Бизнес/Мотивация"],
    ["Iman Gadzhi","@ImanGadzhi",5960000,"Великобритания","","Бизнес/SMMA"],
    ["Ivan AI","@Ivan_Mikhailov_AI",16000,"Индонезия","TG","AI"],
    ["KINZA","@kinzapro",6510,"Кипр","TG,IG,FB,VK","AdTech"],
    ["Kirill Evans","@KirillEvans",434000,"Австралия","TG","Бизнес/Маркетинг"],
    ["KupetsTeam","@kupetsTeam",2180,"Украина","TG,IG","Арбитраж"],
    ["Mark Tilbury","@marktilbury",8200000,"Великобритания","IG,FB","Финансы/Бизнес"],
    ["Mr Mozart","@mr_mozart",2280000,"Россия","TG","Финансы/Бизнес"],
    ["OhMyGod agency","@ohmygodagency1224",619,"Украина","","Маркетинг"],
    ["PIRATE CPA","@piratecpa",2390,"","TG","Арбитраж"],
    ["PowerfulASTE","@SukhanovAlexander",14700,"Россия","TG","AI/Автоматизация"],
    ["PythonToday","@PythonToday",220000,"Россия","TG","Программирование"],
    ["Rush Agency","@Rush_agency",6920,"Россия","TG,IG,FB,VK","SEO/Маркетинг"],
    ["Ruyter Poubel","@RuyterPoubel",3440,"Бразилия","IG","Бизнес"],
    ["Senatorov-LIVE","@artem.senatorov_live",3800,"ОАЭ","TG","SMM/Контент"],
    ["Silicon Valley Girl","@SiliconValleyGirl",1340000,"США","IG","Бизнес/Технологии"],
    ["Smart & Talented","@smart_and_talented",14900,"","TG,IG","HR/Ассистенты"],
    ["SYNTX AI","@syntx_ai",12300,"","","AI/Нейросети"],
    ["TRAFFIC DIVER","@trafficdiver",3790,"Россия","TG","Арбитраж"],
    ["Traffic Light CPA","@trafficlightcpa",2410,"","","Арбитраж"],
    ["Try CGI","@TryCGI",3920,"Таиланд","TG,IG","AI/CGI/Видео"],
    ["Viktor AI","@ViktorrrAI",22100,"Украина","","AI/Автоматизация"],
    ["Web3nity","@Web3nity",658,"","TG","Web3"],
    ["Yoda Traffic","@yodatraffic",9400,"","TG","Арбитраж"],
    ["ZM team","@ZM_team",1630,"","TG","Арбитраж"],
    ["Аве Кодер","@avecoder",62100,"Великобритания","TG,IG,VK","Программирование/AI"],
    ["Александр Высоцкий","@alexvisotsky",622000,"США","TG,IG","Бизнес/Управление"],
    ["Александр Соколовский","@AlexanderSokolovskiy",1350000,"Россия","TG,IG,FB,VK","Бизнес/Маркетинг"],
    ["Александр Сычугов","@sasha_sychugov",2300,"","IG","Путешествия"],
    ["Алексей Ясный","@yasnicreative",3350,"","TG,IG","Креатив"],
    ["Антон Дьяченко","@Anton_Diachenko",16300,"Украина","IG","Бизнес/Инвестиции"],
    ["Антон Клевцов","@klevtsovanton",20100,"Россия","TG","Маркетинг"],
    ["Артур Хорошев","@АртурХорошев",5750,"Россия","TG","AI/Автоматизация"],
    ["Бизнес Стоянов","@BusinessStoyanov",41100,"Россия","TG,VK","Бизнес"],
    ["Бизнес-клуб Атланты","@winningtheheartsAtlantyGroup",16700,"Россия","TG,IG,FB","Нетворкинг"],
    ["Бла Бла Про AI","@BlaBlaProAi",2280,"Россия","TG","AI"],
    ["Богатейший Ди","@TheRichestDee",10200,"Украина","TG","Финансы"],
    ["Взяли Связали","@svyazali",3090,"","","Маркетинг"],
    ["Визави","@vizavi.official",209000,"Казахстан","TG","Подкаст/Бизнес"],
    ["Виктор Ремигайло","@theremigailo",7300,"Россия","TG,IG","Бизнес"],
    ["Владимир Колесов","@vov4ip_smm",161000,"Россия","IG","SMM/Маркетинг"],
    ["Георгий Ривера","@georgiirivera",14600,"Португалия","TG","Бизнес"],
    ["Глеб Соломин","@glebsolomin",1100000,"","","AI/Технологии"],
    ["Денчик Таргетолог","@denkosen",2640,"Россия","TG,IG,VK","Таргетинг"],
    ["Денис Марков","@denismarkov",10200,"Россия","TG","Бизнес"],
    ["Денисенко Денис","@Denisenko_Denis",2650,"","TG","Бизнес"],
    ["Дмитрий Хван","@ДмитрийХван-ГИВОН",18700,"","","Бизнес/Стартапы"],
    ["Дневник Мусульманина","@Mamedblog",486000,"Россия","TG,IG","Лайфстайл"],
    ["ИИ работает за тебя","@AI_to_business",17700,"Казахстан","TG","AI/Бизнес"],
    ["ИИШНЫЙ","@aishny",14500,"","TG","AI"],
    ["Иван AI на практике","@ivan_ai_practice",6370,"Сербия","TG","AI"],
    ["Игнат Егоров AppBusters","@sprestay",17900,"Россия","TG,IG,VK","AI/Приложения"],
    ["Капиталифорния","@Capitalifornia",49400,"США","TG","Бизнес/Инвестиции"],
    ["Кирилл Комаров","@kirillkomarov_",15300,"","","Бизнес"],
    ["Лысый из браузера","@baldfrombrowser",9790,"Казахстан","TG","AI/Автоматизация"],
    ["Люди PRO","@peoplepro",891000,"США","TG,IG,FB","Бизнес/Интервью"],
    ["Макс Афанасьев","@max_afanasjev",8320,"","","Бизнес"],
    ["Максим Морев","@MorevOnline",861,"","TG","Бизнес"],
    ["Максим Новиков","@maksnovikovv",807,"","","Бизнес"],
    ["Маргарита Захарова","@marketing_dir",32500,"Казахстан","TG,IG","Маркетинг"],
    ["Михаил Гребенюк","@grebenukm",516000,"Россия","TG","Бизнес/Продажи"],
    ["Михаил Дашкиев","@dashkiev.m",107000,"","IG,VK","Бизнес"],
    ["Михаил Тимочко","@mikhail.timochko",99700,"Россия","TG","Бизнес/Маркетинг"],
    ["Ната Анарбаева","@anarbacho",31800,"","TG","Маркетинг"],
    ["Наука Соломина","@solomin_science",62100,"ОАЭ","TG","Наука/Образование"],
    ["Нейропросвещение","@neuropros",11600,"Грузия","TG","AI/Нейросети"],
    ["Никита Ефимов","@yefimov_AI",8080,"","TG,IG","AI/Автоматизация"],
    ["Никита Райков","@raykov_nikita",43000,"Таиланд","TG,IG","Маркетинг/Бизнес"],
    ["Оскар Хартманн","@oskar_hartmann1",585000,"Россия","TG,IG,FB,VK","Бизнес/Инвестиции"],
    ["Павел Радкевич","@Pasha_radkevich",2770,"","TG","AI/Маркетинг"],
    ["Павел Гительман","@gitelman",906000,"Россия","TG,IG","Маркетинг/Реклама"],
    ["Павел Гительман Бизнес","@potok_gitelman",60300,"ОАЭ","TG,IG","Бизнес"],
    ["Партнерский маркетинг","@affpodcasts",19300,"США","","Affiliate/Подкаст"],
    ["Практический арбитраж","@leadgenerals",8120,"","TG","Арбитраж"],
    ["РАД","@rad.gilmanov",101000,"","TG,IG","Бизнес/Маркетинг"],
    ["РАДКАСТ","@РАДКАСТ",3780,"","TG","Подкаст"],
    ["Рома Райт","@romarayt",68600,"Грузия","TG","AI/Контент"],
    ["Седа Каспарова","@SedaKasparova",1060000,"ОАЭ","TG,IG","Бизнес/Маркетинг"],
    ["Тимофей Skynetics","@Тимофей_Скайнетикс",18100,"","TG","AI/Автоматизация"],
    ["Трое про ROI","@troeproroi",9060,"Кипр","TG,IG","Маркетинг/ROI"],
    ["Усатый Арбитражник","@usaffiliate",21800,"Украина","TG","Арбитраж"],
    ["Хедлайнеры Ануфриев","@AnufrievNikita",328000,"ОАЭ","","Бизнес/Подкаст"],
    ["Эльнур Аббасов","@e-ducatenow",25600,"Азербайджан","TG","Арбитраж/Media Buying"],
    ["Юлия Трус","@truspro",359000,"Россия","TG","Бизнес/Маркетинг"],
]

# ============================================================
# ВСТРОЕННЫЕ ДАННЫЕ: Telegram (60 ссылок)
# ============================================================
TELEGRAM_DATA = [
    {"handle": "grebenukm", "name": "Михаил Гребенюк", "subscribers": 271603, "description": "Предприниматель, Бизнес-ресурс «Аномалия», Медицинский family-office «ЕС-клиника»"},
    {"handle": "syntxfamily", "name": "SYNTX FAMILY", "subscribers": 320261, "description": "Канал для тех, кто строит будущее с нейросетями. syntx.ai"},
    {"handle": "aiseven_webinar", "name": "Контент-завод + Автоматизация", "subscribers": 6144, "description": "Бесплатные эфиры на тему ИИ: вирусный контент, ИИ-клоны, автоматизация продаж"},
    {"handle": "cgevent", "name": "Метаверсище и ИИще", "subscribers": 50036, "description": "Персональный экспертный взгляд на ИИ, графику, метаверс, крипту, нейротехнологии"},
    {"handle": "aiffring", "name": "Ффринг | Вайбкодинг", "subscribers": 17398, "description": "Строю AI-агентов. Вайбкодинг, автоматизация, уроки и кейсы. Основатель aizavod.io, plaan.ai"},
    {"handle": "pishchulov", "name": "Пищулов Роман (про IT, Ai, FinTech)", "subscribers": 6511, "description": "Тренды на рынке ИИ, CEO SETUP, фаундер школы ИИ-визация"},
    {"handle": "ai_scince", "name": "ananian.ai", "subscribers": 21161, "description": "Делаю Ai-агентов и автоматизации, создаю контент с помощью Ai"},
    {"handle": "senatorovgpt", "name": "Senatorov-GPT", "subscribers": 4621, "description": "Нейросети, AI, SMM и прочее"},
    {"handle": "neyr0graph", "name": "NeuroGraph", "subscribers": 34204, "description": "Midjourney, Seedance, VEO, Nano Banana, ИИ, генерация фото и видео"},
    {"handle": "serge_ai", "name": "Serge_AI 1.0", "subscribers": 61177, "description": ""},
    {"handle": "autogenerator_reels", "name": "Денис про Контент-завод | Автоматизация | Вайбкодинг", "subscribers": 6994, "description": "ТГ-бот по автогенерации Reels, Tik-Tok, Shorts, ИИ-модели"},
    {"handle": "EgorKuzminxr1", "name": "Егор Кузьмин XR", "subscribers": 18974, "description": "Создаю контент с помощью ИИ и преподаю это"},
    {"handle": "cocopolictg", "name": "cocopolic про нейросети", "subscribers": 8090, "description": "AI контент: фото и видео генерация, Midjourney, Seedance, Kling AI"},
    {"handle": "baldfrombrowser", "name": "Лысый из браузера", "subscribers": 16955, "description": "Артём Исламов. Про работу, нейросети, бизнес и жизнь"},
    {"handle": "mirneyrosetey", "name": "Илья Палий", "subscribers": 32840, "description": ""},
    {"handle": "neironick", "name": "Никита о Нейросетях", "subscribers": 2190, "description": "Head of AI Business Transformation Department"},
    {"handle": "bond_nastyaa", "name": "BOND o НЕЙРОСЕТЯХ И МАРКЕТИНГЕ", "subscribers": 16005, "description": "Продажи через контент и нейросети"},
    {"handle": "kirillbezikov", "name": "Нейро Ликбезик | Кирилл Безиков", "subscribers": 19414, "description": "Автоматизация бизнеса при помощи ИИ"},
    {"handle": "viktorrrai", "name": "VIKTOR AI - NEUROHACKER", "subscribers": 2843, "description": "ИИ и автоматизация, обучение"},
    {"handle": "gumirovbrothers", "name": "Братья Гумировы", "subscribers": 8266, "description": "Делаем 1,5-2 млн руб в месяц используя ИИ"},
    {"handle": "yefimov_ai", "name": "Yefimov AI Автоматизация & Бизнес", "subscribers": 2422, "description": "Автоматизации и разработка"},
    {"handle": "breakout_ai", "name": "Breakout AI: все про ИИ, ChatGPT", "subscribers": 21772, "description": "Новые фишки ИИ"},
    {"handle": "maya_pro", "name": "Ковчег / автоматизация make.com и агенты", "subscribers": 11943, "description": "Автоматизация с make, агентами и нейросетями"},
    {"handle": "jedaiit", "name": "Цифровой Джедай", "subscribers": 4236, "description": "Pinterest Маркетинг, Нейросети, бесплатный трафик"},
    {"handle": "generalov_ai_agents", "name": "GENERALOV AI про ИИ и Автоматизацию", "subscribers": 47400, "description": "Обучение пяти ИИ направлениям"},
    {"handle": "artamonov_proreels", "name": "Кирилл Артамонов", "subscribers": 24766, "description": "Бренд-продюсер, эксперт по трафику, личные бренды"},
    {"handle": "masterskaya_video", "name": "AI Мастерская Александра Фисенкова", "subscribers": 8307, "description": "ChatGPT, Midjourney, Sora, RunWay, Kling, туториалы"},
    {"handle": "web3nity_channel", "name": "Hanna Berji | Web3nity", "subscribers": 43757, "description": "Мысли, Развитие, Бизнес, ИИ. email: annletscollab@gmail.com"},
    {"handle": "vizavi_academy", "name": "VIZAVI ACADEMY", "subscribers": 10252, "description": "YouTube Production, ИИ, Онлайн Бизнес"},
    {"handle": "pavel_radkevch", "name": "Павел | Продвижение с ИИ", "subscribers": 1863, "description": "Бизнес по продвижению с нейросетями, ai-zavod.pro"},
    {"handle": "chatplaceio", "name": "ChatPlace", "subscribers": 44259, "description": "Официальный канал сервиса chatplace.io"},
    {"handle": "ArinaAlexx", "name": "Арина Алекс. REELS и продажи", "subscribers": 91537, "description": "Тренды, идеи reels, фишки блогинга и продаж"},
    {"handle": "oshestakovdigital", "name": "Oleg Shestakov - Digital Business", "subscribers": 11426, "description": "Digital бизнес, маркетинг, SEO"},
    {"handle": "contentproduser", "name": "Контент-продюсеры с SKR", "subscribers": 867, "description": "Площадка для заказчиков и продюсеров"},
    {"handle": "marketing_dir", "name": "Подруга маркетолог (Маргарита Захарова)", "subscribers": 22004, "description": "Контент маркетинг, инструменты, разборы"},
    {"handle": "sokolay", "name": "Sokolovskiy", "subscribers": 112828, "description": "Предприниматель, бизнес-подкаст №1"},
    {"handle": "nechto_community", "name": "Нечто", "subscribers": 23264, "description": "Бизнес сообщество для предпринимателей, топ-менеджеров"},
    {"handle": "metodkomarova", "name": "Кирилл Комаров", "subscribers": 4532, "description": "Кейсы, маркетинг"},
    {"handle": "youtube_marina", "name": "кабинет маркетолога", "subscribers": 52133, "description": "Международное креаторское сообщество Марины Могилко"},
    {"handle": "mtimochko", "name": "Миша Тимочко", "subscribers": 62532, "description": ""},
    {"handle": "nastiayoutubeinsider", "name": "YouTube Insider", "subscribers": 2125, "description": "Ex-сотрудница YouTube про алгоритмы платформы"},
    {"handle": "dimatorgov", "name": "Дмитрий Торгов", "subscribers": 11940, "description": "Бизнес, бренд-маркетинг, сооснователь ChatPlace.io"},
    {"handle": "Mr_ReelsRoyce", "name": "Джафар | Миллионы на контенте", "subscribers": 11076, "description": "Reels на миллионы просмотров, кейсы, тренды"},
    {"handle": "potok_gitelman", "name": "Павел Гительман", "subscribers": 28470, "description": "С командой RTA создал млрд. бизнес. Новый бизнес сеть падел-клубов Buenos Padel"},
    {"handle": "akravtsovcom", "name": "Александр Кравцов | Следующий уровень", "subscribers": 34872, "description": "Александр Кравцов. Старший товарищ."},
    {"handle": "assistantexpress", "name": "ВАКАНСИИ ОНЛАЙН-АССИСТЕНТОВ", "subscribers": 3164, "description": "Сервис от компании Smart&Talented. Канал вакансий удаленных ассистентов"},
    {"handle": "MargulanSeissembai", "name": "Маргулан Сейсембай", "subscribers": 214842, "description": "Исследователь, бизнесмен, инвестор, путешественник"},
    {"handle": "egorovavtelegram", "name": "ГЛАВНЫЕ ПО АССИСТЕНТАМ", "subscribers": 1809, "description": "Академия ассистентов и агентство по подбору Smart & Talented. Более 9 лет на рынке"},
    {"handle": "ArbitrazhTrafficUA", "name": "Арбитраж Трафика чат UA", "subscribers": 4960, "description": "Чат для общения арбитражников"},
    {"handle": "cpa_mediabuyers", "name": "Чат медиабайеров | Арбитраж трафика", "subscribers": 18888, "description": "Чат байеров для обмена опытом и общения. cpalenta.ru"},
    {"handle": "glazcukera", "name": "Глаз Цукера", "subscribers": 7781, "description": "Мануалы, гайды по арбитражу трафика и CPA маркетингу"},
    {"handle": "razzboyniki", "name": "Эльнур Аббасов", "subscribers": 4931, "description": "Автор канала - открыт новым идеям и рад сотрудничеству"},
    {"handle": "Arbitrazh_vakansii_rabota", "name": "Арбитраж. Все вакансии в одном месте. LENKEP", "subscribers": 22827, "description": "Вакансии для специалистов по арбитражу в одном месте"},
    {"handle": "partnerkin_job", "name": "Вакансии: Арбитраж трафика, SEO, Affiliate [Партнеркин]", "subscribers": 8735, "description": "Лучшие вакансии арбитража трафика, CPA и SEO. partnerkin.com"},
    {"handle": "kashjob", "name": "KashJob | Вакансии Арбитраж", "subscribers": 3248, "description": "Площадка найма в арбитраже и digital. kashjob.pro"},
    {"handle": "arbihunter", "name": "ArbiHunter - работа в Affiliate Marketing", "subscribers": 4916, "description": "HR-платформа по подбору персонала в Affiliate marketing, Crypto, Ai и Digital"},
    {"handle": "+3AIcbvomlFNlMDRl", "name": "Приватный канал", "subscribers": 0, "description": ""},
    {"handle": "+vf04teyf8Pc0OTky", "name": "Приватный канал", "subscribers": 0, "description": ""},
    {"handle": "+G-l6aXgllb8wOTA6", "name": "Приватный канал", "subscribers": 0, "description": ""},
]

# ============================================================
# КЛЮЧЕВЫЕ СЛОВА ДЛЯ ОПРЕДЕЛЕНИЯ РОЛЕЙ
# ============================================================
ROLE_KEYWORDS = {
    "Инвестор": ["инвест", "invest", "капитал", "capital", "фонд", "fund", "venture", "vc ", "angel", "бизнес-ангел"],
    "AI-автоматизатор": ["ai ", "нейро", "neuro", "автоматиз", "automat", "gpt", "chatgpt", "нейросет", "neural", "machine learning", "ml ", "искусственн"],
    "Контент-мейкер": ["контент", "content", "reels", "видео", "video", "creator", "мейкер", "maker", "продакшн", "production", "монтаж"],
    "Блогер": ["блог", "blog", "влог", "vlog", "youtube", "тикток", "tiktok", "инстаграм", "подкаст", "podcast"],
    "Бизнес": ["бизнес", "business", "предпринимат", "entrepreneur", "ceo", "founder", "основатель", "стартап", "startup", "компани"],
    "Маркетолог": ["маркет", "market", "smm", "реклам", "advertis", "таргет", "target", "трафик", "traffic", "продвиж", "промо"],
    "Affiliate": ["affiliate", "аффилиат", "партнёр", "cpa", "арбитраж", "arbitr", "офферы", "media buy", "gambling"],
    "Образование": ["образован", "educat", "обучен", "курс", "course", "школ", "school", "академи", "academ", "тренинг", "менторств"],
    "Недвижимость": ["недвижим", "real estate", "риелтор", "realtor", "квартир", "жильё", "property"],
    "Финансы/Крипто": ["финанс", "financ", "крипто", "crypto", "трейд", "trade", "биткоин", "bitcoin", "деньг", "money", "forex", "web3"],
}


# ============================================================
# ФУНКЦИИ
# ============================================================

def parse_number(val):
    """Парсинг числа подписчиков: '3177651' -> 3177651, '8.2M' -> 8200000"""
    if not val:
        return 0
    val = str(val).strip().replace(" ", "").replace(",", "").replace("\xa0", "")
    try:
        return int(val)
    except ValueError:
        pass
    # Обработка суффиксов K, M
    m = re.match(r'^([\d.]+)\s*([KkМмMm]?)$', val)
    if m:
        num = float(m.group(1))
        suffix = m.group(2).upper()
        if suffix in ('K', 'К'):
            return int(num * 1000)
        elif suffix in ('M', 'М'):
            return int(num * 1000000)
        return int(num)
    try:
        return int(float(val))
    except:
        return 0


def detect_roles(text):
    """Определяет роли на основе ключевых слов в тексте"""
    if not text:
        return []
    text_lower = text.lower()
    roles = []
    for role, keywords in ROLE_KEYWORDS.items():
        for kw in keywords:
            if kw in text_lower:
                roles.append(role)
                break
    return roles


def calc_priority(followers, verified, has_email, is_business, roles):
    """
    Приоритет A/B/C:
    A = верифицирован + 500K+ ИЛИ email + 100K+
    B = 50K+ ИЛИ email ИЛИ бизнес + 20K+
    C = остальные
    """
    if (verified and followers >= 500000) or (has_email and followers >= 100000):
        return "A"
    if followers >= 50000 or has_email or (is_business and followers >= 20000):
        return "B"
    return "C"


def normalize_handle(handle):
    """Нормализация хендла для сравнения"""
    h = str(handle).lower().strip()
    h = h.lstrip("@")
    h = h.replace("_", "").replace(".", "").replace("-", "")
    return h


def levenshtein(s1, s2):
    """Расстояние Левенштейна"""
    if len(s1) < len(s2):
        return levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        curr_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = prev_row[j + 1] + 1
            deletions = curr_row[j] + 1
            subs = prev_row[j] + (c1 != c2)
            curr_row.append(min(insertions, deletions, subs))
        prev_row = curr_row
    return prev_row[-1]


def read_instagram_csv(filepath):
    """Читает Instagram CSV и возвращает список словарей"""
    rows = []
    seen = set()
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            if len(row) < 13:
                row.extend([""] * (13 - len(row)))
            username = row[0].strip()
            if not username:
                continue
            if username.lower() in DUPLICATES:
                continue
            if username.lower() in seen:
                continue
            seen.add(username.lower())

            followers = parse_number(row[3])
            is_verified = row[8].strip().lower() in ("да", "yes", "true", "1")
            is_business = row[7].strip().lower() in ("да", "yes", "true", "1")
            email = row[10].strip()
            bio = row[9].strip() if len(row) > 9 else ""
            category = row[6].strip() if len(row) > 6 else ""
            website = row[11].strip() if len(row) > 11 else ""
            extra_links = row[12].strip() if len(row) > 12 else ""

            combined_text = f"{bio} {category} {website} {extra_links}"
            roles = detect_roles(combined_text)
            priority = calc_priority(followers, is_verified, bool(email), is_business, roles)

            rows.append({
                "username": username,
                "name": row[1].strip(),
                "url": row[2].strip(),
                "followers": followers,
                "following": parse_number(row[4]),
                "posts": parse_number(row[5]),
                "category": category,
                "is_business": "Да" if is_business else "Нет",
                "is_verified": "Да" if is_verified else "Нет",
                "bio": bio,
                "email": email,
                "website": website,
                "extra_links": extra_links,
                "roles": ", ".join(roles) if roles else "—",
                "priority": priority,
            })
    return rows


def process_youtube():
    """Обрабатывает YouTube данные"""
    rows = []
    for yt in YOUTUBE_DATA:
        name, handle, subs, country, socials, topic = yt
        subs_num = parse_number(subs)
        combined = f"{name} {topic} {socials}"
        roles = detect_roles(combined)

        # Приоритет для YouTube
        has_tg = "TG" in str(socials)
        has_ig = "IG" in str(socials)
        if subs_num >= 500000 or (subs_num >= 100000 and (has_tg or has_ig)):
            pri = "A"
        elif subs_num >= 50000 or has_tg or has_ig:
            pri = "B"
        else:
            pri = "C"

        rows.append({
            "name": name,
            "handle": handle,
            "subscribers": subs_num,
            "country": country,
            "socials": socials,
            "topic": topic,
            "roles": ", ".join(roles) if roles else "—",
            "priority": pri,
        })
    return rows


def process_telegram(ig_data, yt_data):
    """Обрабатывает Telegram данные и матчит с IG/YT"""
    rows = []
    for tg in TELEGRAM_DATA:
        handle = tg["handle"]
        is_private = handle.startswith("+")
        tg_url = f"https://t.me/{handle}"
        tg_norm = normalize_handle(handle)

        # Поиск матча с Instagram
        ig_match = ""
        for ig in ig_data:
            ig_norm = normalize_handle(ig["username"])
            if tg_norm == ig_norm or levenshtein(tg_norm, ig_norm) <= 2:
                ig_match = ig["username"]
                break
            # Проверка в доп. ссылках
            if handle in str(ig.get("extra_links", "")):
                ig_match = ig["username"]
                break

        # Поиск матча с YouTube
        yt_match = ""
        for yt in yt_data:
            yt_norm = normalize_handle(yt["handle"])
            if tg_norm == yt_norm or levenshtein(tg_norm, yt_norm) <= 2:
                yt_match = yt["name"]
                break

        rows.append({
            "handle": handle,
            "name": tg.get("name", ""),
            "subscribers": tg.get("subscribers", 0),
            "description": tg.get("description", ""),
            "url": tg_url,
            "private": "Да" if is_private else "Нет",
            "ig_match": ig_match,
            "yt_match": yt_match,
        })
    return rows


def find_crossplatform(ig_data, yt_data, tg_data):
    """Находит людей, которые есть на нескольких платформах"""
    cross = []
    for ig in ig_data:
        ig_norm = normalize_handle(ig["username"])
        platforms = ["Instagram"]
        yt_name = ""
        tg_handle = ""

        for yt in yt_data:
            yt_norm = normalize_handle(yt["handle"])
            if ig_norm == yt_norm or levenshtein(ig_norm, yt_norm) <= 2:
                platforms.append("YouTube")
                yt_name = yt["name"]
                break
            # Проверка имён
            if ig["name"] and yt["name"]:
                ig_name_norm = normalize_handle(ig["name"])
                yt_name_norm = normalize_handle(yt["name"])
                if ig_name_norm == yt_name_norm and len(ig_name_norm) > 3:
                    platforms.append("YouTube")
                    yt_name = yt["name"]
                    break

        for tg in tg_data:
            tg_norm = normalize_handle(tg["handle"])
            if ig_norm == tg_norm or levenshtein(ig_norm, tg_norm) <= 2:
                platforms.append("Telegram")
                tg_handle = tg["handle"]
                break

        if len(platforms) >= 2:
            cross.append({
                "name": ig["name"],
                "ig_username": ig["username"],
                "yt_name": yt_name,
                "tg_handle": tg_handle,
                "platforms": ", ".join(platforms),
                "count": len(platforms),
                "ig_followers": ig["followers"],
                "priority": ig["priority"],
                "roles": ig["roles"],
            })
    return sorted(cross, key=lambda x: -x["count"])


def build_outreach(ig_data, yt_data, tg_data):
    """
    Строит единый мастер-лист: один человек = одна строка.
    Объединяет данные со всех платформ, исключает дубли.
    """
    people = []
    used_ig = set()
    used_yt = set()
    used_tg = set()

    # --- Шаг 1: Начинаем с Instagram, ищем матчи с YT и TG ---
    for ig in ig_data:
        ig_norm = normalize_handle(ig["username"])
        p = {
            "name": ig["name"] or ig["username"],
            # Instagram
            "ig_user": ig["username"],
            "ig_url": f'https://instagram.com/{ig["username"]}',
            "ig_followers": ig["followers"],
            "ig_posts": ig["posts"],
            "ig_bio": ig["bio"],
            "ig_category": ig["category"],
            "ig_verified": ig["is_verified"],
            "ig_business": ig["is_business"],
            "ig_email": ig["email"],
            "ig_website": ig["website"],
            "ig_extra_links": ig["extra_links"],
            "ig_roles": ig["roles"],
            # YouTube (пустые по умолчанию)
            "yt_name": "", "yt_handle": "", "yt_url": "",
            "yt_subs": 0, "yt_country": "", "yt_topic": "", "yt_socials": "",
            # Telegram (пустые по умолчанию)
            "tg_handle": "", "tg_url": "", "tg_subs": 0, "tg_desc": "",
            # Мета
            "platforms": 1, "platform_list": ["IG"],
        }

        # Поиск матча с YouTube
        for j, yt in enumerate(yt_data):
            if j in used_yt:
                continue
            yt_norm = normalize_handle(yt["handle"])
            matched = (ig_norm == yt_norm or levenshtein(ig_norm, yt_norm) <= 2)
            if not matched and ig["name"] and yt["name"]:
                in_ = normalize_handle(ig["name"])
                yn_ = normalize_handle(yt["name"])
                if in_ == yn_ and len(in_) > 3:
                    matched = True
            if matched:
                p["yt_name"] = yt["name"]
                p["yt_handle"] = yt["handle"]
                p["yt_url"] = f'https://youtube.com/{yt["handle"]}'
                p["yt_subs"] = yt["subscribers"]
                p["yt_country"] = yt["country"]
                p["yt_topic"] = yt["topic"]
                p["yt_socials"] = yt["socials"]
                p["platforms"] += 1
                p["platform_list"].append("YT")
                used_yt.add(j)
                break

        # Поиск матча с Telegram
        for k, tg in enumerate(tg_data):
            if k in used_tg:
                continue
            tg_norm = normalize_handle(tg["handle"])
            matched = (ig_norm == tg_norm or levenshtein(ig_norm, tg_norm) <= 2)
            if not matched and tg["handle"] in str(ig.get("extra_links", "")):
                matched = True
            if matched:
                p["tg_handle"] = tg["handle"]
                p["tg_url"] = f'https://t.me/{tg["handle"]}'
                p["tg_subs"] = tg["subscribers"]
                p["tg_desc"] = tg.get("description", "")
                p["platforms"] += 1
                p["platform_list"].append("TG")
                used_tg.add(k)
                break

        used_ig.add(ig["username"])
        people.append(p)

    # --- Шаг 2: YouTube без матча с IG, ищем TG ---
    for j, yt in enumerate(yt_data):
        if j in used_yt:
            continue
        yt_norm = normalize_handle(yt["handle"])
        p = {
            "name": yt["name"],
            "ig_user": "", "ig_url": "", "ig_followers": 0, "ig_posts": 0,
            "ig_bio": "", "ig_category": "", "ig_verified": "", "ig_business": "",
            "ig_email": "", "ig_website": "", "ig_extra_links": "", "ig_roles": "",
            "yt_name": yt["name"], "yt_handle": yt["handle"],
            "yt_url": f'https://youtube.com/{yt["handle"]}',
            "yt_subs": yt["subscribers"], "yt_country": yt["country"],
            "yt_topic": yt["topic"], "yt_socials": yt["socials"],
            "tg_handle": "", "tg_url": "", "tg_subs": 0, "tg_desc": "",
            "platforms": 1, "platform_list": ["YT"],
        }

        for k, tg in enumerate(tg_data):
            if k in used_tg:
                continue
            tg_norm = normalize_handle(tg["handle"])
            if yt_norm == tg_norm or levenshtein(yt_norm, tg_norm) <= 2:
                p["tg_handle"] = tg["handle"]
                p["tg_url"] = f'https://t.me/{tg["handle"]}'
                p["tg_subs"] = tg["subscribers"]
                p["tg_desc"] = tg.get("description", "")
                p["platforms"] += 1
                p["platform_list"].append("TG")
                used_tg.add(k)
                break

        used_yt.add(j)
        people.append(p)

    # --- Шаг 3: Telegram без матча с IG/YT (пропускаем приватные) ---
    for k, tg in enumerate(tg_data):
        if k in used_tg:
            continue
        if tg["handle"].startswith("+"):
            continue
        p = {
            "name": tg.get("name", "") or tg["handle"],
            "ig_user": "", "ig_url": "", "ig_followers": 0, "ig_posts": 0,
            "ig_bio": "", "ig_category": "", "ig_verified": "", "ig_business": "",
            "ig_email": "", "ig_website": "", "ig_extra_links": "", "ig_roles": "",
            "yt_name": "", "yt_handle": "", "yt_url": "",
            "yt_subs": 0, "yt_country": "", "yt_topic": "", "yt_socials": "",
            "tg_handle": tg["handle"],
            "tg_url": f'https://t.me/{tg["handle"]}',
            "tg_subs": tg["subscribers"],
            "tg_desc": tg.get("description", ""),
            "platforms": 1, "platform_list": ["TG"],
        }
        used_tg.add(k)
        people.append(p)

    # --- Шаг 4: Рассчитываем приоритет и собираем мета ---
    RELEVANT_ROLES = {"AI-автоматизатор", "Контент-мейкер", "Affiliate", "Маркетолог"}
    for p in people:
        total = p["ig_followers"] + p["yt_subs"] + p["tg_subs"]
        p["total_followers"] = total

        # Объединённые роли
        all_text = " ".join(filter(None, [
            p["ig_bio"], p["ig_category"], p["ig_roles"],
            p["yt_topic"], p["tg_desc"],
        ]))
        roles = detect_roles(all_text)
        p["roles"] = ", ".join(roles) if roles else "—"
        is_relevant = bool(set(roles) & RELEVANT_ROLES)

        # Приоритет
        plat = p["platforms"]
        if plat >= 3:
            p["priority"] = "A"
        elif plat >= 2 and (total >= 100000 or is_relevant):
            p["priority"] = "A"
        elif total >= 500000:
            p["priority"] = "A"
        elif is_relevant and total >= 50000:
            p["priority"] = "A"
        elif plat >= 2 or is_relevant or total >= 20000 or p["ig_email"]:
            p["priority"] = "B"
        else:
            p["priority"] = "C"

        # Описание (лучшее из доступных)
        p["description"] = p["ig_bio"] or p["tg_desc"] or p["yt_topic"] or ""
        # Категория
        p["category"] = p["ig_category"] or p["yt_topic"] or ""
        # Email
        p["email"] = p["ig_email"] or ""
        # Сайт
        p["website"] = p["ig_website"] or ""
        # Доп. ссылки
        p["extra_links"] = p["ig_extra_links"] or ""

    # Сортировка: платформы ↓, приоритет A>B>C, подписчики ↓
    pri_order = {"A": 0, "B": 1, "C": 2}
    people.sort(key=lambda x: (
        -x["platforms"],
        pri_order.get(x["priority"], 3),
        -x["total_followers"],
    ))

    return people


# ============================================================
# СТИЛИ EXCEL
# ============================================================
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

PRIORITY_A_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
PRIORITY_B_FILL = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
PRIORITY_C_FILL = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")

CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
NUMBER_ALIGN = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

STAT_LABEL_FONT = Font(name="Arial", bold=True, size=12)
STAT_VALUE_FONT = Font(name="Arial", size=12, color="DAA520")


def style_header(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(cell, is_number=False):
    cell.font = CELL_FONT
    cell.alignment = NUMBER_ALIGN if is_number else CELL_ALIGN
    cell.border = THIN_BORDER


def style_priority(cell, priority):
    if priority == "A":
        cell.fill = PRIORITY_A_FILL
        cell.font = Font(name="Arial", bold=True, size=10)
    elif priority == "B":
        cell.fill = PRIORITY_B_FILL
    elif priority == "C":
        cell.fill = PRIORITY_C_FILL


def auto_width(ws, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        adjusted = min(max_len + 2, max_width)
        ws.column_dimensions[col_letter].width = max(adjusted, 10)


# ============================================================
# СОЗДАНИЕ EXCEL
# ============================================================
def create_excel(ig_data, yt_data, tg_data, cross_data):
    wb = openpyxl.Workbook()

    # ======= ЛИСТ 1: Instagram =======
    ws_ig = wb.active
    ws_ig.title = "Instagram"
    ig_headers = ["Username", "Имя", "URL", "Подписчики", "Подписки", "Публикации",
                  "Категория", "Бизнес", "Верифицирован", "Bio", "Email", "Сайт",
                  "Доп. ссылки", "Роли", "Приоритет"]
    ws_ig.append(ig_headers)
    style_header(ws_ig, 1, len(ig_headers))

    for i, ig in enumerate(ig_data, 2):
        ws_ig.append([
            ig["username"], ig["name"], ig["url"], ig["followers"],
            ig["following"], ig["posts"], ig["category"], ig["is_business"],
            ig["is_verified"], ig["bio"][:200], ig["email"], ig["website"],
            ig["extra_links"][:200], ig["roles"], ig["priority"]
        ])
        for col in range(1, len(ig_headers) + 1):
            cell = ws_ig.cell(row=i, column=col)
            style_data_cell(cell, col in (4, 5, 6))
        style_priority(ws_ig.cell(row=i, column=15), ig["priority"])

    ws_ig.auto_filter.ref = f"A1:O{len(ig_data)+1}"
    ws_ig.freeze_panes = "A2"
    auto_width(ws_ig)

    # ======= ЛИСТ 2: YouTube =======
    ws_yt = wb.create_sheet("YouTube")
    yt_headers = ["Название", "Handle", "Подписчики", "Страна", "Соцсети",
                  "Тематика", "Роли", "Приоритет"]
    ws_yt.append(yt_headers)
    style_header(ws_yt, 1, len(yt_headers))

    for i, yt in enumerate(yt_data, 2):
        ws_yt.append([
            yt["name"], yt["handle"], yt["subscribers"], yt["country"],
            yt["socials"], yt["topic"], yt["roles"], yt["priority"]
        ])
        for col in range(1, len(yt_headers) + 1):
            cell = ws_yt.cell(row=i, column=col)
            style_data_cell(cell, col == 3)
        style_priority(ws_yt.cell(row=i, column=8), yt["priority"])

    ws_yt.auto_filter.ref = f"A1:H{len(yt_data)+1}"
    ws_yt.freeze_panes = "A2"
    auto_width(ws_yt)

    # ======= ЛИСТ 3: Telegram =======
    ws_tg = wb.create_sheet("Telegram")
    tg_headers = ["Handle", "Название", "Подписчики", "Описание", "URL", "Приватный", "Матч Instagram", "Матч YouTube"]
    ws_tg.append(tg_headers)
    style_header(ws_tg, 1, len(tg_headers))

    for i, tg in enumerate(tg_data, 2):
        ws_tg.append([
            tg["handle"], tg["name"], tg["subscribers"] if tg["subscribers"] else "",
            tg["description"], tg["url"], tg["private"], tg["ig_match"], tg["yt_match"]
        ])
        for col in range(1, len(tg_headers) + 1):
            style_data_cell(ws_tg.cell(row=i, column=col))

    ws_tg.auto_filter.ref = f"A1:H{len(tg_data)+1}"
    ws_tg.freeze_panes = "A2"
    auto_width(ws_tg)

    # ======= ЛИСТ 4: Приоритет A =======
    ws_a = wb.create_sheet("Приоритет A")
    a_headers = ["Платформа", "Имя/Название", "Username/Handle", "Подписчики",
                 "Роли", "Email", "Приоритет"]
    ws_a.append(a_headers)
    style_header(ws_a, 1, len(a_headers))

    row_num = 2
    for ig in ig_data:
        if ig["priority"] == "A":
            ws_a.append(["Instagram", ig["name"], ig["username"], ig["followers"],
                         ig["roles"], ig["email"], "A"])
            for col in range(1, len(a_headers) + 1):
                cell = ws_a.cell(row=row_num, column=col)
                style_data_cell(cell, col == 4)
            style_priority(ws_a.cell(row=row_num, column=7), "A")
            row_num += 1

    for yt in yt_data:
        if yt["priority"] == "A":
            ws_a.append(["YouTube", yt["name"], yt["handle"], yt["subscribers"],
                         yt["roles"], "", "A"])
            for col in range(1, len(a_headers) + 1):
                cell = ws_a.cell(row=row_num, column=col)
                style_data_cell(cell, col == 4)
            style_priority(ws_a.cell(row=row_num, column=7), "A")
            row_num += 1

    ws_a.auto_filter.ref = f"A1:G{row_num-1}"
    ws_a.freeze_panes = "A2"
    auto_width(ws_a)

    # ======= ЛИСТ 5: Кросс-платформа =======
    ws_cross = wb.create_sheet("Кросс-платформа")
    cross_headers = ["Имя", "Instagram", "YouTube", "Telegram", "Платформы",
                     "Кол-во", "Подписчики IG", "Приоритет", "Роли"]
    ws_cross.append(cross_headers)
    style_header(ws_cross, 1, len(cross_headers))

    for i, cr in enumerate(cross_data, 2):
        ws_cross.append([
            cr["name"], cr["ig_username"], cr["yt_name"], cr["tg_handle"],
            cr["platforms"], cr["count"], cr["ig_followers"], cr["priority"], cr["roles"]
        ])
        for col in range(1, len(cross_headers) + 1):
            cell = ws_cross.cell(row=i, column=col)
            style_data_cell(cell, col in (6, 7))
        style_priority(ws_cross.cell(row=i, column=8), cr["priority"])

    ws_cross.auto_filter.ref = f"A1:I{len(cross_data)+1}"
    ws_cross.freeze_panes = "A2"
    auto_width(ws_cross)

    # ======= ЛИСТ 6: С Email =======
    ws_email = wb.create_sheet("С Email")
    email_headers = ["Платформа", "Имя", "Username", "Email", "Подписчики",
                     "Роли", "Приоритет"]
    ws_email.append(email_headers)
    style_header(ws_email, 1, len(email_headers))

    row_num = 2
    for ig in ig_data:
        if ig["email"]:
            ws_email.append(["Instagram", ig["name"], ig["username"], ig["email"],
                             ig["followers"], ig["roles"], ig["priority"]])
            for col in range(1, len(email_headers) + 1):
                cell = ws_email.cell(row=row_num, column=col)
                style_data_cell(cell, col == 5)
            style_priority(ws_email.cell(row=row_num, column=7), ig["priority"])
            row_num += 1

    ws_email.auto_filter.ref = f"A1:G{row_num-1}"
    ws_email.freeze_panes = "A2"
    auto_width(ws_email)

    # ======= ЛИСТ 7: АУТРИЧ (мастер-лист) =======
    outreach = build_outreach(ig_data, yt_data, tg_data)
    ws_out = wb.create_sheet("Аутрич")
    out_headers = [
        "№", "Приоритет", "Имя", "Платформы", "Роли", "Категория",
        "Описание",
        "YouTube", "YT подписчики",
        "Instagram", "IG подписчики", "IG публикации",
        "Telegram", "TG подписчики",
        "Email", "Сайт", "Доп. ссылки",
        "Всего подписчиков", "Страна",
        "Статус", "Дата контакта", "Примечания",
    ]
    ws_out.append(out_headers)
    style_header(ws_out, 1, len(out_headers))

    # Секции-разделители
    section_rows = {}  # row_num -> label
    row_num = 2
    prev_plat = None
    for idx, p in enumerate(outreach):
        plat = p["platforms"]
        if plat != prev_plat:
            # Вставляем разделитель секции
            label = {3: "3 ПЛАТФОРМЫ", 2: "2 ПЛАТФОРМЫ", 1: "1 ПЛАТФОРМА"}.get(plat, "")
            ws_out.cell(row=row_num, column=1, value=label)
            ws_out.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(out_headers))
            section_cell = ws_out.cell(row=row_num, column=1)
            section_cell.font = Font(name="Arial", bold=True, size=11, color="DAA520")
            section_cell.fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
            section_cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1
            prev_plat = plat

        yt_link = p["yt_url"] if p["yt_url"] else ""
        ig_link = p["ig_url"] if p["ig_url"] else ""
        tg_link = p["tg_url"] if p["tg_url"] else ""

        row_data = [
            idx + 1,
            p["priority"],
            p["name"],
            p["platforms"],
            p["roles"],
            p["category"],
            p["description"][:300],
            yt_link,
            p["yt_subs"] if p["yt_subs"] else "",
            ig_link,
            p["ig_followers"] if p["ig_followers"] else "",
            p["ig_posts"] if p["ig_posts"] else "",
            tg_link,
            p["tg_subs"] if p["tg_subs"] else "",
            p["email"],
            p["website"],
            p["extra_links"][:200],
            p["total_followers"],
            p.get("yt_country", ""),
            "",  # Статус
            "",  # Дата контакта
            "",  # Примечания
        ]
        ws_out.append(row_data)

        for col in range(1, len(out_headers) + 1):
            cell = ws_out.cell(row=row_num, column=col)
            style_data_cell(cell, col in (4, 9, 11, 12, 14, 18))
            # Делаем ссылки кликабельными
            if col == 8 and yt_link:
                cell.hyperlink = yt_link
                cell.font = Font(name="Arial", size=10, color="4472C4", underline="single")
            elif col == 10 and ig_link:
                cell.hyperlink = ig_link
                cell.font = Font(name="Arial", size=10, color="E1306C", underline="single")
            elif col == 13 and tg_link:
                cell.hyperlink = tg_link
                cell.font = Font(name="Arial", size=10, color="0088CC", underline="single")
            elif col == 15 and p["email"]:
                cell.hyperlink = f'mailto:{p["email"]}'
                cell.font = Font(name="Arial", size=10, color="4472C4", underline="single")

        style_priority(ws_out.cell(row=row_num, column=2), p["priority"])
        row_num += 1

    ws_out.auto_filter.ref = f"A1:V{row_num-1}"
    ws_out.freeze_panes = "D2"

    # Ширина колонок
    out_widths = {
        "A": 5, "B": 10, "C": 25, "D": 10, "E": 25, "F": 20,
        "G": 40, "H": 30, "I": 14, "J": 30, "K": 14, "L": 12,
        "M": 25, "N": 14, "O": 25, "P": 25, "Q": 25, "R": 16,
        "S": 12, "T": 14, "U": 14, "V": 25,
    }
    for col_letter, w in out_widths.items():
        ws_out.column_dimensions[col_letter].width = w

    # ======= ЛИСТ 8: Статистика =======
    ws_stat = wb.create_sheet("Статистика")

    ig_a = sum(1 for x in ig_data if x["priority"] == "A")
    ig_b = sum(1 for x in ig_data if x["priority"] == "B")
    ig_c = sum(1 for x in ig_data if x["priority"] == "C")
    yt_a = sum(1 for x in yt_data if x["priority"] == "A")
    yt_b = sum(1 for x in yt_data if x["priority"] == "B")
    yt_c = sum(1 for x in yt_data if x["priority"] == "C")
    emails_count = sum(1 for x in ig_data if x["email"])
    cross_count = len(cross_data)

    # Подсчёт ролей
    role_counter = defaultdict(int)
    for ig in ig_data:
        for r in ig["roles"].split(", "):
            if r != "—":
                role_counter[r] += 1
    for yt in yt_data:
        for r in yt["roles"].split(", "):
            if r != "—":
                role_counter[r] += 1

    stats = [
        ["ОБЩАЯ СТАТИСТИКА CRM", ""],
        ["", ""],
        ["INSTAGRAM", ""],
        ["Всего профилей (после дедупликации)", len(ig_data)],
        ["Удалено дубликатов", len(DUPLICATES)],
        ["Приоритет A", ig_a],
        ["Приоритет B", ig_b],
        ["Приоритет C", ig_c],
        ["С email", emails_count],
        ["Верифицированных", sum(1 for x in ig_data if x["is_verified"] == "Да")],
        ["Бизнес-аккаунтов", sum(1 for x in ig_data if x["is_business"] == "Да")],
        ["", ""],
        ["YOUTUBE", ""],
        ["Всего каналов", len(yt_data)],
        ["Приоритет A", yt_a],
        ["Приоритет B", yt_b],
        ["Приоритет C", yt_c],
        ["", ""],
        ["TELEGRAM", ""],
        ["Всего ссылок", len(TELEGRAM_DATA)],
        ["Приватных", sum(1 for t in TELEGRAM_DATA if t["handle"].startswith("+"))],
        ["Публичных", sum(1 for t in TELEGRAM_DATA if not t["handle"].startswith("+"))],
        ["", ""],
        ["КРОСС-ПЛАТФОРМА", ""],
        ["Людей на 2+ платформах", cross_count],
        ["", ""],
        ["РОЛИ (все платформы)", ""],
    ]

    for role, count in sorted(role_counter.items(), key=lambda x: -x[1]):
        stats.append([f"  {role}", count])

    # Статистика Аутрич
    out_a = sum(1 for x in outreach if x["priority"] == "A")
    out_b = sum(1 for x in outreach if x["priority"] == "B")
    out_c = sum(1 for x in outreach if x["priority"] == "C")
    out_3 = sum(1 for x in outreach if x["platforms"] == 3)
    out_2 = sum(1 for x in outreach if x["platforms"] == 2)
    out_1 = sum(1 for x in outreach if x["platforms"] == 1)

    stats.extend([
        ["", ""],
        ["АУТРИЧ (мастер-лист)", ""],
        ["Уникальных людей", len(outreach)],
        ["На 3 платформах", out_3],
        ["На 2 платформах", out_2],
        ["На 1 платформе", out_1],
        ["Приоритет A", out_a],
        ["Приоритет B", out_b],
        ["Приоритет C", out_c],
        ["", ""],
        ["ИТОГО", ""],
        ["Всего контактов (с дублями)", len(ig_data) + len(yt_data) + len(TELEGRAM_DATA)],
        ["Уникальных людей (без дублей)", len(outreach)],
        ["Приоритет A (все)", out_a],
        ["Приоритет B (все)", out_b],
        [f"Дата генерации", datetime.now().strftime("%d.%m.%Y %H:%M")],
    ])

    for row in stats:
        ws_stat.append(row)

    # Стили для статистики
    for i in range(1, ws_stat.max_row + 1):
        cell_a = ws_stat.cell(row=i, column=1)
        cell_b = ws_stat.cell(row=i, column=2)
        if cell_a.value and str(cell_a.value).isupper() and cell_a.value not in ("", None):
            cell_a.font = Font(name="Arial", bold=True, size=12, color="DAA520")
        else:
            cell_a.font = STAT_LABEL_FONT
        cell_b.font = STAT_VALUE_FONT
        cell_b.alignment = Alignment(horizontal="right")

    ws_stat.column_dimensions["A"].width = 45
    ws_stat.column_dimensions["B"].width = 20

    return wb


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("  CRM Generator — Shikardos Team")
    print("=" * 60)

    # 1. Instagram
    if not os.path.exists(IG_CSV):
        print(f"\n❌ Файл не найден: {IG_CSV}")
        print("Убедись, что CRM_Instagram_DeepScan_442.csv лежит рядом со скриптом.")
        sys.exit(1)

    print(f"\n📥 Читаю Instagram CSV: {IG_CSV}")
    ig_data = read_instagram_csv(IG_CSV)
    print(f"   ✅ {len(ig_data)} профилей (после удаления {len(DUPLICATES)} дубликатов)")

    # 2. YouTube
    print(f"\n📺 Обработка YouTube: {len(YOUTUBE_DATA)} каналов")
    yt_data = process_youtube()
    print(f"   ✅ {len(yt_data)} каналов обработано")

    # 3. Telegram
    print(f"\n✈️ Обработка Telegram: {len(TELEGRAM_DATA)} ссылок")
    tg_data = process_telegram(ig_data, yt_data)
    print(f"   ✅ {len(tg_data)} ссылок обработано")

    # 4. Кросс-платформа
    print(f"\n🔗 Поиск кросс-платформенных совпадений...")
    cross_data = find_crossplatform(ig_data, yt_data, tg_data)
    print(f"   ✅ {len(cross_data)} совпадений найдено")

    # 5. Создание Excel
    print(f"\n📊 Создаю Excel файл...")
    wb = create_excel(ig_data, yt_data, tg_data, cross_data)

    wb.save(OUTPUT)
    print(f"\n✅ ГОТОВО! Файл сохранён: {OUTPUT}")
    print(f"   Листов: 8")
    print(f"   Instagram: {len(ig_data)} строк")
    print(f"   YouTube: {len(yt_data)} строк")
    print(f"   Telegram: {len(tg_data)} строк")
    print(f"   Приоритет A: {sum(1 for x in ig_data if x['priority'] == 'A') + sum(1 for x in yt_data if x['priority'] == 'A')} контактов")
    print(f"   Кросс-платформа: {len(cross_data)} совпадений")
    print(f"   С Email: {sum(1 for x in ig_data if x['email'])} контактов")
    print(f"\n🎯 АУТРИЧ (мастер-лист):")

    # Подсчёт аутрич-статистики
    outreach = build_outreach(ig_data, yt_data, tg_data)
    out_3 = sum(1 for x in outreach if x["platforms"] == 3)
    out_2 = sum(1 for x in outreach if x["platforms"] == 2)
    out_1 = sum(1 for x in outreach if x["platforms"] == 1)
    out_a = sum(1 for x in outreach if x["priority"] == "A")
    print(f"   Уникальных людей: {len(outreach)}")
    print(f"   На 3 платформах: {out_3}")
    print(f"   На 2 платформах: {out_2}")
    print(f"   На 1 платформе: {out_1}")
    print(f"   Приоритет A: {out_a}")

    print(f"\n📂 Открой файл: {OUTPUT}")


if __name__ == "__main__":
    main()
